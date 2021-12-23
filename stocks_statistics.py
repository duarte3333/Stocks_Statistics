import datetime
import pandas as pd
import pandas_datareader.data as web
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from prettytable import PrettyTable 


class Stock:
    def __init__(self, nome, precoCompra , precoAtual,Euro,dia,mes,ano):
        self.n = nome
        self.c = precoCompra
        if ("." in nome):
            self.a = precoAtual
        
        else: 
            self.a = precoAtual*(1/Euro)
       

    def get_nome(self):
        return self.n
    
    def get_precoCompra(self):
        return self.c
    
    def get_precoAtual(self):
        return self.a
    
    def get_profit(self):
        self.p = self.a - self.c
        return self.p

def number_acoes():
    number = 21
    return number

#funcao que le ficheiros excel e recebe data de compra e valor da compra
def excel_reader():

    # Give the location of the file
    #loc = ("Users\Duarte Morais\Desktop")
    workbook = load_workbook(filename="acoes.xlsx")
    sheet = workbook.active
    #a = sheet["A1"].value
    f = sheet.iter_rows(min_row = 2,max_row = number_acoes() ,min_col = 1, max_col = 3)
    ano = []
    mes = []
    dia = []
    datas_datetime = []
    
    preco_compra = []
    names = []
    for a,b,c in f:
        
        datas = a.value
        datas_datetime.append(datas)
        ano.append(datas.year)
        mes.append(datas.month)
        dia.append(datas.day)
        
        preco_compra.append(b.value)
        names.append(c.value)
    print(names)
    print(preco_compra)
    return dia,mes,ano,preco_compra,names,datas_datetime

#funcao que cria uma acao e dá append na lista de acoes
def criar_acao(dia,mes,ano,preco_compra,nomes):
    
    start = datetime.datetime(2020,10, 10)
    end = datetime.datetime.now()
    EURUSD = web.DataReader("EURUSD=X",'yahoo', start, end)
    Stocks = []
    
    print("Aguarde um pouco...")
  
    for i in range(len(dia)):
    
        name = web.DataReader(nomes[i],'yahoo', start, end)
        Stock_atual = Stock(nomes[i],preco_compra[i],name.Close[-1],EURUSD.Close[-1],dia[i],mes[i],ano[i])
        #print(Stock_atual)
        Stocks.append(Stock_atual)
    return Stocks


#funcao que faz o relatorio dizendo(racio,total investido,profit,top5 melhores investimentos)
def estatisticas(ListaStocks):
    
    total_inv = 0
    for a in range(0,len(ListaStocks)):
        total_inv = total_inv + ListaStocks[a].get_precoCompra()

    print("Total invested: " + str(total_inv) + "€")
    
    profitT = 0
    lista_profts = []
    lista_profits_acumulados = []
    
    for a in range(0,len(ListaStocks)):
        profitT = profitT + ListaStocks[a].get_profit()
        lista_profts.append(ListaStocks[a].get_profit())
        lista_profits_acumulados.append(profitT)
        
        
    print("Total profit: " + str(profitT.round(3)) + "€")
    
    ratio = profitT/total_inv*100
    print("Total ratio: " + str(ratio.round(3)) + "%")
    
    return lista_profits_acumulados

def Sort_Tuple(tup): 
      
    # getting length of list of tuples
    lst = len(tup) 
    for i in range(0, lst): 
          
        for j in range(0, lst-i-1): 
            if (tup[j][1] > tup[j + 1][1]): 
                temp = tup[j] 
                tup[j]= tup[j + 1] 
                tup[j + 1]= temp 
    return tup     

def top5_investimentos(ListaStocks):
    
    profits = []
    names = []
    
    for Stock in ListaStocks:
        
        Stock.get_nome()
        Stock.get_profit()
        profits.append(Stock.get_profit())
        names.append(Stock.get_nome())

    
    tuple_list = []

    for m in range(len(profits)):
        matched_indexes = []
        i = 0
        while i < len(names):
            if names[m] == names[i]:
                matched_indexes.append(i)
            i += 1
            
        total = 0
        for a in range(0,len(matched_indexes)):
            
            b = matched_indexes[a]
            total = total + profits[b]
        
        tuple1 = (names[m],total)
        tuple_list.append(tuple1)
        
        tuple_dudes = list(dict.fromkeys(tuple_list))
        mytuples = Sort_Tuple(tuple_dudes)
    
    myTable = PrettyTable(["Ranking", "Name" ,"Profit"]) 
  
    # Add rows 
    myTable.add_row(["1",str(mytuples[-1][0]), str(mytuples[-1][1])]) 
    myTable.add_row(["2",str(mytuples[-2][0]), str(mytuples[-2][1])]) 
    myTable.add_row(["3",str(mytuples[-3][0]), str(mytuples[-3][1])]) 
    myTable.add_row(["4",str(mytuples[-4][0]), str(mytuples[-4][1])]) 
    myTable.add_row(["5",str(mytuples[-5][0]), str(mytuples[-5][1])]) 
  
      
    print(myTable)
  
    
    return profits
      
#funcao que cria um gráfico a escolha de duas variaveis
def criar_grafico(x, y ,title_r, x_label, y_label):
    
    ts = pd.Series(y, x)
    
    close = ts
    ax = close.plot(title=str(title_r))
    ax.set_xlabel(x_label)
    ax.set_ylabel(y_label)
    ax.grid()
    plt.show()

def saber_percentagens(names,preco_compra):
    
    total_inv = sum(preco_compra)
    
    tuple_list = []
    for m in range(len(preco_compra)):
        matched_indexes = []
        i = 0
        while i < len(preco_compra):
            if names[m] == names[i]:
                matched_indexes.append(i)
            i += 1
            
        total = 0
        for a in range(0,len(matched_indexes)):
            
            b = matched_indexes[a]
            total = total + preco_compra[b]
        final = total/total_inv * 100
        tuple1 = (names[m],final)
        tuple_list.append(tuple1)
        
        tuple_dudes = list(dict.fromkeys(tuple_list))
        
    #print(mytuples)
    return tuple_dudes
    
    
def func(pct, allvalues):
    
    absolute = int(pct / 100.*np.sum(allvalues))
   
    return "{:.1f}%\n".format(pct, absolute)

#Esta funcao só funciona se manualmente for colocado as percentagens e os nomes das acoes
def graph_circular(p):
    
    # Creating dataset
    
    #stocks1 = ["SPX-500", 'AIandRoboticsETF','Nvidea', 'Semiconductors', 'AMD',
    #       'Palantir', 'NIO', 'C3AI','Ai']
      
    #data1 = [ 46,2, 18, 16, 6, 2, 3, 5,2]
    
    stocks = []
    data = []
    for i in range(len(p)):
        stocks.append(str(p[i][0]))
        data.append(p[i][1])
      
    
    #ATENCAO: NUMERO DE ENTRADAS NO EXPLODE E NO COLORS TEM DE TER O MESMO NUMERO DE 
    #ENTRADAS QUE A LISTA CARS
    # Creating explode data
    explode = (0.1, 0.0, 0.2, 0.1, 0.2, 0.0,0.1,0.1,0.2,0.2)
      
    # Creating color parameters
    colors = ( "orange", "cyan", "brown",
              "grey", "purple", "beige","blue","green","red","yellow")
      
    # Wedge properties
    wp = { 'linewidth' : 1, 'edgecolor' : "black" }
    
    fig, ax = plt.subplots(figsize =(10, 8))
    wedges, texts, autotexts = ax.pie(data, 
                                      autopct = lambda pct: func(pct, data),
                                      explode = explode, 
                                      labels = stocks,
                                      shadow = True,
                                      colors = colors,
                                      startangle = 90,
                                      wedgeprops = wp,
                                      textprops = dict(color ="black"))
      
    # Adding legend
    ax.legend(wedges, stocks,
              title ="Stocks",
              loc ="upper right",
              bbox_to_anchor =(1, 0, 0.5, 1))
      
    plt.setp(autotexts, size = 10, weight ="bold")
    ax.set_title("Duarte Stock's",size=16,loc = "left")
      
    # show plot
    plt.show()


#criar funçao que sugere compra vender ou manter
def main():
    dia, mes, ano, preco_compra , nomes, datas = excel_reader()
    
    percentagens = saber_percentagens(nomes,preco_compra)
    
    lista_acoes = criar_acao(dia, mes, ano, preco_compra, nomes)
    lista_profits_acumulados = estatisticas(lista_acoes)
    top5_investimentos(lista_acoes)
    
    #Plot do dinheiro investido ao longo do tempo
    inv_acumuladoF = []
    total = 0
    for i in range(len(preco_compra)):
        
        total = total + preco_compra[i]
        inv_acumuladoF.append(total)
        
    #inv_acumulado = [52.91,118.53,218.64,277.96,308.84,440.79,532.07,599.39,885.42,1085.95,1163.11]
    
    #mylist = list(dict.fromkeys(datas))
    
    criar_grafico(datas, inv_acumuladoF,"Gráfico: Total dinheiro investido - Tempo","Tempo", "Dinheiro Investido total")
    criar_grafico(datas, lista_profits_acumulados,"Grafico: Profit total - Tempo","Tempo", "Profit total")
    graph_circular(percentagens)
    
main()




  





